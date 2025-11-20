from fastapi import FastAPI, HTTPException
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.middleware.cors import CORSMiddleware
import requests
import json
import os
import tempfile
from datetime import datetime
from typing import Optional, List
from pydantic import BaseModel
import openpyxl
from openpyxl.styles import Font, Alignment
from dotenv import load_dotenv

# Load environment variables
load_dotenv()

# --- API Models ---

class JobSearchRequest(BaseModel):
    """Request model for job search parameters"""
    query: str
    num_pages: Optional[int] = 2
    country: Optional[str] = "us"
    date_posted: Optional[str] = "all"
    page: Optional[int] = 1
    language: Optional[str] = None
    work_from_home: Optional[bool] = False
    employment_types: Optional[str] = None
    job_requirements: Optional[str] = None
    radius: Optional[int] = None
    exclude_job_publishers: Optional[str] = None
    fields: Optional[str] = None

# --- API Configuration ---

app = FastAPI(
    title="Job Search API",
    description="API for searching jobs and generating formatted HTML results",
    version="1.0.0"
)

# API Configuration from environment variables
RAPIDAPI_KEY = os.getenv('RAPIDAPI_KEY', '3733bc90bdmsh3a4e3a23d0b9629p179849jsnbf97a06e60ed')
API_URL = 'https://jsearch.p.rapidapi.com/search'

# CORS Configuration
CORS_ALLOWED_ORIGINS = os.getenv('CORS_ALLOWED_ORIGINS')
if CORS_ALLOWED_ORIGINS:
    allowed_origins = [origin.strip() for origin in CORS_ALLOWED_ORIGINS.split(',')]
else:
    allowed_origins = ["*"]

# Add CORS middleware
app.add_middleware(
    CORSMiddleware,
    allow_origins=allowed_origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# --- XLSX Generation Functions ---

def generate_xlsx_filename(query: str) -> str:
    """Generate filename for XLSX file using first 12 chars of query and current date"""
    # Get first 12 characters of query, remove special characters
    query_part = query[:12].replace(' ', '_').replace('/', '_').replace('\\', '_')
    # Get current date in day-month-year format
    current_date = datetime.now().strftime("%d-%m-%Y")
    return f"{query_part}-{current_date}.xlsx"

def create_xlsx_file(job_list: List[dict], query: str) -> str:
    """Create XLSX file with ALL job data from RapidAPI response"""
    print(f"Creating XLSX with {len(job_list)} jobs")
    
    # DEBUG: Check what we're actually getting
    print(f"🔍 DEBUG: job_list type: {type(job_list)}")
    if job_list:
        print(f"🔍 DEBUG: First job type: {type(job_list[0])}")
        print(f"🔍 DEBUG: First job content sample: {str(job_list[0])[:200]}...")
    
    # Create workbook and worksheet
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Job Search Results"
    
    # Define comprehensive column headers based on ALL available data from RapidAPI
    headers = [
        "Job ID", "Job Title", "Employer Name", "Employer Logo", "Employer Website", 
        "Employer Company Type", "Employer LinkedIn", "Job Publisher", 
        "Employment Type", "Employment Types", "Employment Type Text",
        "Apply Link", "Apply Is Direct", "Apply Quality Score",
        "Job Description", "Is Remote", "Posted Human Readable", 
        "Posted Timestamp", "Posted Datetime UTC", "Location", "City", 
        "State", "Country", "Latitude", "Longitude", "Benefits",
        "Google Link", "Offer Expiration UTC", "Offer Expiration Timestamp",
        "Required Experience", "No Experience Required", "Required Experience Months",
        "Experience Mentioned", "Experience Preferred", "Salary", "Min Salary", 
        "Max Salary", "Salary Currency", "Salary Period", "Highlights",
        "Job Job Title", "Posting Language", "ONET SOC", "ONET Job Zone",
        "Occupational Categories", "NAICS Code", "NAICS Name"
    ]
    
    # Add header row with styling
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Add job data rows - extracting ALL fields from RapidAPI response
    for row, job in enumerate(job_list, 2):
        # Extract ALL available data fields from the API response
        worksheet.cell(row=row, column=1, value=job.get('job_id', 'N/A'))
        worksheet.cell(row=row, column=2, value=job.get('job_title', 'N/A'))
        worksheet.cell(row=row, column=3, value=job.get('employer_name', 'N/A'))
        worksheet.cell(row=row, column=4, value=job.get('employer_logo', 'N/A'))
        worksheet.cell(row=row, column=5, value=job.get('employer_website', 'N/A'))
        worksheet.cell(row=row, column=6, value=job.get('employer_company_type', 'N/A'))
        worksheet.cell(row=row, column=7, value=job.get('employer_linkedin', 'N/A'))
        worksheet.cell(row=row, column=8, value=job.get('job_publisher', 'N/A'))
        worksheet.cell(row=row, column=9, value=job.get('job_employment_type', 'N/A'))
        worksheet.cell(row=row, column=10, value=str(job.get('job_employment_types', [])))
        worksheet.cell(row=row, column=11, value=job.get('job_employment_type_text', 'N/A'))
        worksheet.cell(row=row, column=12, value=job.get('job_apply_link', 'N/A'))
        worksheet.cell(row=row, column=13, value=job.get('job_apply_is_direct', 'N/A'))
        worksheet.cell(row=row, column=14, value=job.get('job_apply_quality_score', 'N/A'))
        worksheet.cell(row=row, column=15, value=job.get('job_description', 'N/A'))
        worksheet.cell(row=row, column=16, value=job.get('job_is_remote', 'N/A'))
        worksheet.cell(row=row, column=17, value=job.get('job_posted_human_readable', 'N/A'))
        worksheet.cell(row=row, column=18, value=job.get('job_posted_at_timestamp', 'N/A'))
        worksheet.cell(row=row, column=19, value=job.get('job_posted_at_datetime_utc', 'N/A'))
        worksheet.cell(row=row, column=20, value=job.get('job_location', 'N/A'))
        worksheet.cell(row=row, column=21, value=job.get('job_city', 'N/A'))
        worksheet.cell(row=row, column=22, value=job.get('job_state', 'N/A'))
        worksheet.cell(row=row, column=23, value=job.get('job_country', 'N/A'))
        worksheet.cell(row=row, column=24, value=job.get('job_latitude', 'N/A'))
        worksheet.cell(row=row, column=25, value=job.get('job_longitude', 'N/A'))
        worksheet.cell(row=row, column=26, value=str(job.get('job_benefits', [])))
        worksheet.cell(row=row, column=27, value=job.get('job_google_link', 'N/A'))
        worksheet.cell(row=row, column=28, value=job.get('job_offer_expiration_datetime_utc', 'N/A'))
        worksheet.cell(row=row, column=29, value=job.get('job_offer_expiration_timestamp', 'N/A'))
        
        # Handle nested required_experience object
        required_exp = job.get('job_required_experience', {})
        worksheet.cell(row=row, column=30, value=str(required_exp))
        worksheet.cell(row=row, column=31, value=required_exp.get('no_experience_required', 'N/A'))
        worksheet.cell(row=row, column=32, value=required_exp.get('required_experience_in_months', 'N/A'))
        worksheet.cell(row=row, column=33, value=required_exp.get('experience_mentioned', 'N/A'))
        worksheet.cell(row=row, column=34, value=required_exp.get('experience_preferred', 'N/A'))
        
        worksheet.cell(row=row, column=35, value=job.get('job_salary', 'N/A'))
        worksheet.cell(row=row, column=36, value=job.get('job_min_salary', 'N/A'))
        worksheet.cell(row=row, column=37, value=job.get('job_max_salary', 'N/A'))
        worksheet.cell(row=row, column=38, value=job.get('job_salary_currency', 'N/A'))
        worksheet.cell(row=row, column=39, value=job.get('job_salary_period', 'N/A'))
        worksheet.cell(row=row, column=40, value=str(job.get('job_highlights', {})))
        worksheet.cell(row=row, column=41, value=job.get('job_job_title', 'N/A'))
        worksheet.cell(row=row, column=42, value=job.get('job_posting_language', 'N/A'))
        worksheet.cell(row=row, column=43, value=job.get('job_onet_soc', 'N/A'))
        worksheet.cell(row=row, column=44, value=job.get('job_onet_job_zone', 'N/A'))
        worksheet.cell(row=row, column=45, value=job.get('job_occupational_categories', 'N/A'))
        worksheet.cell(row=row, column=46, value=job.get('job_naics_code', 'N/A'))
        worksheet.cell(row=row, column=47, value=job.get('job_naics_name', 'N/A'))
    
    # Auto-adjust column widths for better readability
    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)  # Cap at 30 characters for better readability
        worksheet.column_dimensions[column_letter].width = adjusted_width
    
    # Save to temporary file
    filename = generate_xlsx_filename(query)
    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, filename)
    workbook.save(file_path)
    
    print(f"✅ XLSX created with {len(job_list)} jobs and {len(headers)} data fields")
    return file_path

# --- HTML Template Functions ---

def generate_job_accordion(job_data, index):
    """Generates the HTML for a single, collapsible job entry."""
    job_title = job_data.get('job_title', 'N/A')
    employer_name = job_data.get('employer_name', 'N/A')
    location = job_data.get('job_location', 'N/A')
    employment_type = job_data.get('job_employment_type', 'N/A')
    posted_at = job_data.get('job_posted_at', 'N/A')
    apply_link = job_data.get('job_apply_link')
    description = job_data.get('job_description', 'No detailed description available.').replace('\n', '<br>')
    
    # Format salary if available
    min_salary = job_data.get('job_min_salary')
    max_salary = job_data.get('job_max_salary')
    salary_period = job_data.get('job_salary_period')
    
    salary_str = "N/A"
    if min_salary and max_salary and salary_period:
        salary_str = f"${min_salary:,} - ${max_salary:,} ({salary_period})"
    elif job_data.get('job_salary'):
        salary_str = job_data['job_salary']

    # Use a descriptive link for the apply URL
    apply_link_html = f'<a href="{apply_link}" target="_blank">Apply Here</a>' if apply_link else 'N/A'
    
    # Accordion Structure
    return f"""
    <div class="job-card">
        <input type="checkbox" id="job{index}" class="accordion-toggle">
        <label for="job{index}" class="accordion-header">
            <span class="job-title-label">Job Title:</span>
            <span class="job-title-text">{job_title}</span>
            <span class="location-text"> @ {location}</span>
        </label>
        <div class="accordion-content">
            <p><strong>Employer:</strong> {employer_name}</p>
            <p><strong>Location:</strong> {location}</p>
            <p><strong>Type:</strong> {employment_type}</p>
            <p><strong>Posted:</strong> {posted_at}</p>
            <p><strong>Salary:</strong> {salary_str}</p>
            <p class="apply-link"><strong>Apply Link:</strong> {apply_link_html}</p>
            <div class="description-section">
                <h4>Job Description:</h4>
                <p>{description}</p>
            </div>
        </div>
    </div>
    """

def create_full_html(job_list, query, request_id):
    """Generates the full HTML file content."""
    
    total_jobs = len(job_list)
    job_accordions = "".join([generate_job_accordion(job, i) for i, job in enumerate(job_list)])
    
    # Generate XLSX filename for the export button
    xlsx_filename = generate_xlsx_filename(query)
    
    # CSS for the 'nice easy to read well formatted' style and accordion
    css = """
    body { font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background-color: #f4f7f9; color: #333; margin: 20px; }
    .container { max-width: 900px; margin: 0 auto; background-color: #ffffff; padding: 20px 40px; border-radius: 12px; box-shadow: 0 6px 15px rgba(0, 0, 0, 0.1); }
    h1 { color: #007bff; text-align: center; border-bottom: 3px solid #007bff; padding-bottom: 10px; margin-bottom: 25px; }
    .summary { text-align: center; margin-bottom: 30px; font-size: 1.1em; color: #555; }
    .export-section { text-align: center; margin-bottom: 20px; }
    .export-button { background-color: #28a745; color: white; border: none; padding: 12px 24px; border-radius: 6px; cursor: pointer; font-size: 16px; font-weight: bold; text-decoration: none; display: inline-block; transition: background-color 0.3s; }
    .export-button:hover { background-color: #218838; text-decoration: none; color: white; }
    .job-card { border: 1px solid #ddd; margin-bottom: 10px; border-radius: 8px; overflow: hidden; background-color: #fdfdff; }
    .accordion-toggle { display: none; }
    .accordion-header { display: block; padding: 15px; cursor: pointer; background-color: #e9ecef; color: #333; transition: background-color 0.3s; font-weight: bold; position: relative; }
    .accordion-header:hover { background-color: #d8dde3; }
    .accordion-header::after { content: '\\25B6'; /* Right triangle */ position: absolute; right: 15px; top: 50%; transform: translateY(-50%); transition: transform 0.3s; }
    .accordion-toggle:checked + .accordion-header::after { content: '\\25BC'; /* Down triangle */ }
    .accordion-toggle:checked + .accordion-header { background-color: #007bff; color: white; }
    .accordion-content { max-height: 0; overflow: hidden; transition: max-height 0.3s ease-out, padding 0.3s; padding: 0 15px; }
    .accordion-toggle:checked ~ .accordion-content { max-height: 1000px; /* Large enough value */ padding: 15px; border-top: 1px solid #ddd; }
    .job-title-label { color: #007bff; margin-right: 5px; }
    .job-title-text { color: #333; }
    .location-text { font-size: 0.9em; font-weight: normal; color: #6c757d; }
    .apply-link a { color: #28a745; text-decoration: none; font-weight: bold; }
    .apply-link a:hover { text-decoration: underline; }
    .description-section { margin-top: 15px; border-top: 1px dashed #ccc; padding-top: 10px; }
    .description-section h4 { color: #007bff; margin-top: 0; }
    """
    
    html = f"""
    <!DOCTYPE html>
    <html lang="en">
    <head>
        <meta charset="UTF-8">
        <meta name="viewport" content="width=device-width, initial-scale=1.0">
        <title>Job Search Results: {query}</title>
        <style>{css}</style>
    </head>
    <body>
        <div class="container">
            <h1>🔍 Job Search Results</h1>
            <div class="summary">
                <p><strong>Query:</strong> {query} | <strong>Total Jobs Found:</strong> {total_jobs} | <strong>Request ID:</strong> {request_id}</p>
            </div>
            <div class="export-section">
                <a href="/api/export/xlsx/{request_id}" class="export-button" download="{xlsx_filename}">
                    📊 Export to Excel
                </a>
            </div>
            {job_accordions}
        </div>
    </body>
    </html>
    """
    return html

# Store job data for export (in-memory cache)
job_data_cache = {}

# --- API Endpoints ---

@app.get("/")
async def root():
    """Root endpoint with API information"""
    return {
        "message": "Job Search API",
        "version": "1.0.0",
        "endpoints": {
            "search": "/search (POST) - Search for jobs and get HTML results",
            "docs": "/docs - Interactive API documentation"
        }
    }

@app.post("/search", response_class=HTMLResponse)
async def search_jobs(request: JobSearchRequest):
    """
    Search for jobs and return formatted HTML results
    
    Parameters:
    - query: Search query string (required)
    - num_pages: Number of pages to fetch (default: 2)
    - country: Country code (default: "us")
    - date_posted: Date posted filter (default: "all")
    - page: Starting page (default: 1)
    - language: Language code for job postings (optional)
    - work_from_home: Only return remote jobs (default: false)
    - employment_types: Comma-separated employment types (FULLTIME, CONTRACTOR, PARTTIME, INTERN)
    - job_requirements: Comma-separated job requirements (under_3_years_experience, more_than_3_years_experience, no_experience, no_degree)
    - radius: Search radius from location in km (optional)
    - exclude_job_publishers: Comma-separated publishers to exclude (optional)
    - fields: Comma-separated fields to include in response (optional)
    """
    
    print(f"Processing job search request for: '{request.query}'")
    
    # Setup API Request with all parameters
    querystring = {
        "query": request.query,
        "page": str(request.page),
        "num_pages": str(request.num_pages),
        "country": request.country,
        "date_posted": request.date_posted
    }
    
    # Add optional parameters if provided
    if request.language:
        querystring["language"] = request.language
    if request.work_from_home:
        querystring["work_from_home"] = str(request.work_from_home).lower()
    if request.employment_types:
        querystring["employment_types"] = request.employment_types
    if request.job_requirements:
        querystring["job_requirements"] = request.job_requirements
    if request.radius:
        querystring["radius"] = str(request.radius)
    if request.exclude_job_publishers:
        querystring["exclude_job_publishers"] = request.exclude_job_publishers
    if request.fields:
        querystring["fields"] = request.fields
    
    headers = {
        "x-rapidapi-key": RAPIDAPI_KEY,
        "x-rapidapi-host": "jsearch.p.rapidapi.com"
    }
    
    try:
        # Make API Call
        response = requests.get(API_URL, headers=headers, params=querystring, timeout=15)
        response.raise_for_status()
        data = response.json()
        
    except requests.exceptions.RequestException as e:
        raise HTTPException(status_code=500, detail=f"Error fetching data from API: {str(e)}")

    # Process Data
    job_data_list = data.get('data', [])
    request_id = data.get('request_id', 'unknown-id')
    
    if not job_data_list:
        raise HTTPException(status_code=404, detail="No jobs found for the given search criteria")

    # GENERATE XLSX FILE IMMEDIATELY and store file path in cache
    # This ensures the export link works even if cache is cleared
    try:
        xlsx_file_path = create_xlsx_file(job_data_list, request.query)
        print(f"📊 Generated XLSX file: {xlsx_file_path}")
    except Exception as e:
        print(f"❌ Error generating XLSX: {str(e)}")
        xlsx_file_path = None

    # Store both job data AND XLSX file path in cache
    job_data_cache[request_id] = {
        'job_list': job_data_list,  # This is the actual job data from the API
        'query': request.query,
        'xlsx_file_path': xlsx_file_path  # Store the actual file path
    }

    print(f"💾 Stored {len(job_data_list)} jobs in cache with request_id: {request_id}")
    if job_data_list:
        print(f"📊 Sample cached job keys: {list(job_data_list[0].keys())}")

    # Generate HTML - now the export link will work immediately
    html_content = create_full_html(job_data_list, request.query, request_id)
    
    print(f"✅ Success! Generated HTML for {len(job_data_list)} jobs")
    
    return HTMLResponse(content=html_content)

@app.get("/export/xlsx/{request_id}")
async def export_jobs_xlsx(request_id: str):
    """
    Export job search results to XLSX format
    
    Parameters:
    - request_id: The request ID from the job search
    """
    
    print(f"🔍 Export request received for request_id: {request_id}")
    
    # Check if job data exists for this request_id
    if request_id not in job_data_cache:
        print(f"❌ No data found in cache for request_id: {request_id}")
        raise HTTPException(status_code=404, detail="Job data not found for this request ID")
    
    cached_data = job_data_cache[request_id]
    query = cached_data['query']
    xlsx_file_path = cached_data.get('xlsx_file_path')
    
    print(f"📊 Retrieved export data for query: '{query}'")
    print(f"📊 XLSX file path: {xlsx_file_path}")
    
    # Check if we have a pre-generated XLSX file
    if xlsx_file_path and os.path.exists(xlsx_file_path):
        print(f"✅ Using pre-generated XLSX file: {xlsx_file_path}")
        
        # Generate filename for download
        filename = generate_xlsx_filename(query)
        
        # Return the pre-generated file as download
        return FileResponse(
            path=xlsx_file_path,
            filename=filename,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    else:
        print(f"❌ No pre-generated XLSX file found, generating new one")
        
        # Fallback: generate XLSX file from cached job data
        job_list = cached_data['job_list']
        
        try:
            # Create XLSX file
            file_path = create_xlsx_file(job_list, query)
            
            # Generate filename for download
            filename = generate_xlsx_filename(query)
            
            print(f"✅ Success! Generated XLSX file: {filename}")
            
            # Return file as download
            return FileResponse(
                path=file_path,
                filename=filename,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
        except Exception as e:
            print(f"❌ Error generating XLSX: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error generating XLSX file: {str(e)}")

if __name__ == "__main__":
    import uvicorn
    port = int(os.getenv("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
